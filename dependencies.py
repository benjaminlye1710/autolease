# -*- coding: utf-8 -*-
"""
Created on Fri Dec 24 09:55:05 2021

@author: AnikaLeeTH
"""


### Import Pacakges
import pandas as pd
import numpy as np
import re
import calendar
import datetime
from dateutil.relativedelta import relativedelta

import openpyxl
from openpyxl import utils
from openpyxl import formatting, styles
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment

import math
from copy import copy
import logging
import os
import time
from collections import OrderedDict


# initialise logger	
logger = logging.getLogger()	
if not(logger.hasHandlers()):	
    logger.addHandler(logging.StreamHandler())
    
### Standard Variables
EXTENSIONS_EXCEL = [".xls", ".xlsx", ".xlsb", ".xlsm"]
EXTENSIONS_TEXT = [".txt", ".csv"]

# Ordered container types
ORDERED_CONTAINER_TYPES = [pd.Series, list, tuple, np.array]

# CALENDAR ABBR TO MONTH
month_abbr_to_number = {
    calendar.month_abbr[i] : i
    for i in range(1, 13)
    }

MONTH_ABBR_LST = list(month_abbr_to_number.keys())

MONTH_ABBR_TO_NUMBER = month_abbr_to_number

MONTH_NAME_TO_NUMBER = {
    calendar.month_name[i] : i
    for i in range(1, 13)
    }

MONTH_LST = list(MONTH_NAME_TO_NUMBER.keys())

### common functions
def get_unique_dict_values(input_dict, assert_one_value = False,	
                           keyname = "Key", valname = "values"):	
    '''	
    input_dict is a dict where values are a list	
                        	
    if assert_one is True:	
        return dict of key to val	
        	
    if assert_one is False	
        return dict of key to unique values	
    '''	
    
    #	
    output_dict = OrderedDict()	
    for k, v in input_dict.items():	
        output_dict[k] = list(set(v))	
        	
    if assert_one_value is False:	
        	
        pass	
        	
    else:	
            	
        for key, values in input_dict.items():	
            	
            if len(values) > 1:	
                error = f"{keyname.capitalize()} = {key} has multiple {valname}: {values}."	
                logger.error(error)	
                raise Exception (error)	
            	
            else:	
                	
                output_dict[key] = values[0]	
    	
    return output_dict	
    	    
    
def assert_one_and_get(inlist, what="items"):
    '''
    This function checks whether there's only one element in 
    the list.   
    '''
    
    if len(inlist) > 1:
        raise Exception ("Multiple %s found: %s." % (what, inlist))
    elif len(inlist) == 0:
        raise Exception ("No %s found." % what)
    else:
        return inlist[0]


def shift_month(s, delta=0, month_end=None, month_begin=None,
                parse_fmt=None, strft_fmt=None):
    '''
    adjust month by number of months
    
    s => can be a date, or a string e.g. "feb 2020"
    delta => number of months (negative to reverse)
    parse_fmt => optional, else use ensure_correct_date_format to check date format
    strft_fmt = None -> return date
                else -> return string formatted date
    dayfirst =  True -> assume day first
                False -> assume month first
                
    Example:
    adjust_month('jan 2020', +1) => "Feb 2020"
    adjust_month('jan 2020', -3) => 'Oct 2019'
    adjust_month('jan 2020', +1, strft_fmt=None) = datetime.date(2019, 10, 1)
    '''
    
    # Convert to date
    try:
        d = pd.to_datetime(s, format = parse_fmt, dayfirst = True).date()
    except:
        error = "Cannot parse date for '%s'." % s
        # logger.error(error)
        raise Exception(error)
    
    # Set time unit
    month_unit = relativedelta(months=delta)
    
    # return
    adj_d = d + month_unit
    
    # adjust month_end or month_begin
    if month_end:
        adj_d_day = calendar.monthrange(adj_d.year, adj_d.month)[1]
        adj_d = datetime.date(adj_d.year, adj_d.month, adj_d_day)
    elif month_begin:
        adj_d_day = 1
        adj_d = datetime.date(adj_d.year, adj_d.month, adj_d_day)
    elif month_begin is not None or month_begin is not None: 
        msg = 'Please ensure either month_begin or month_end is True.'
        raise NotImplementedError(msg)
    
    if strft_fmt is None:
        return adj_d
    else:
        return adj_d.strftime(strft_fmt)
    

def ensure_correct_date_format(date_series, date_format = 'dmy', 
                               delimiter = None):
    '''
    This method checks for date format.
    
    date_series => can be a string or pandas.Series
    
    date_format = dmy => dd.mm.yyyy
                = mdy => mm.dd.yyyy
                = ymd => yyyy.mm.dd
                where:
                    d is between 1 to 31
                    m is between 1 to 12, Jan to Dec or January to December
                    y is between 0 to 9999
                    . is the delimiter
                    
    delimiter => a str. 
                ['-', '/', '.', '<space>'] will be used if not specified
    '''
    
    MONTH_ABBR = list(MONTH_ABBR_TO_NUMBER.keys())
    MONTH_NAME = list(MONTH_NAME_TO_NUMBER.keys())
    regex_mth = '|'.join(MONTH_ABBR) + '|' + '|'.join(MONTH_NAME)
    
    if type(date_series) is str:
        date_series = pd.Series(date_series, name = 'Date')
    
    if date_series.dtype == object:
        
        # Drop na and dups
        unique_date_series = date_series.dropna().drop_duplicates()
        unique_date_series.index = unique_date_series.values
        
        # Split into 5 columns
        # Y/M/D/B/b smth Y/M/D/B/b smth Y/M/D/B/b
        if delimiter is None:
            regex_pat = (
                r'(?i)(' + regex_mth +
                r'|[0-9]{1,4})[\-/. ]?(' + regex_mth +
                r'|[0-9]{1,4})[\-/. ]?([0-9]{1,4})'
                )
            
            yr_mth_day_df = \
                unique_date_series.astype(str).str.extract(regex_pat)
        
        else:
            yr_mth_day_df = \
                unique_date_series.str.split(delimiter, expand = True)
            yr_mth_day_df[2] = yr_mth_day_df[2].str.extract('^(\d{1,4})')
                
        if date_format.strip().lower() == 'ymd':
            yr_mth_day_df.rename(columns = {0:'yr', 1:'mth', 2:'day'},
                                 inplace = True, errors = 'raise')
        elif date_format.strip().lower() =='mdy':
            yr_mth_day_df.rename(columns = {0:'mth', 1:'day', 2:'yr'},
                                 inplace = True, errors = 'raise')
        elif date_format.strip().lower() == 'dmy':
            yr_mth_day_df.rename(columns = {0:'day', 1:'mth', 2:'yr'},
                                 inplace = True, errors = 'raise')
        else:
            msg = ('date_format is not ["dmy", "mdy", "dmy"], where '
                   'd = Day, m = Month, y = Year.'
                   )
            # logger.error(msg)
            raise NotImplementedError (msg)
        
        # convert to month name/abbr to number
        yr_mth_day_df['mth'] = yr_mth_day_df['mth'].str.title()
        all_month_name_ind = yr_mth_day_df['mth'].isin(MONTH_NAME)
        all_month_abbr_ind =yr_mth_day_df['mth'].isin(MONTH_ABBR)
        
        if all_month_abbr_ind.any():
            yr_mth_day_df['mth'] = yr_mth_day_df['mth'].replace(MONTH_ABBR_TO_NUMBER)
        
        elif all_month_name_ind.any():
            yr_mth_day_df['mth'] = yr_mth_day_df['mth'].replace(MONTH_NAME_TO_NUMBER)
        
        
        day_is_mth_ind = yr_mth_day_df['day'].isin(MONTH_NAME + MONTH_ABBR)
        col_id = yr_mth_day_df.columns.get_loc('day')
        
        if day_is_mth_ind.all() and col_id==1:
            msg = (f'date_format = "{date_format}" is incorrect. It should be '
                   'either ["dmy", "ymd"].'
                   )
            # logger.error(msg)
            raise Exception (msg)
        
        elif day_is_mth_ind.all() and col_id==0:
            msg = (f'date_format = "{date_format}" is incorrect. It should be '
                   '["mdy"].'
                   )
            # logger.error(msg)
            raise Exception (msg)
            
        
        # convert to int
        yr_mth_day_df = yr_mth_day_df.apply(pd.to_numeric, errors ='raise') 
        
        yr_mth_day_df['Day Criterion'] = \
            (yr_mth_day_df['day']>=1) & (yr_mth_day_df['day']<=31)
        
        yr_mth_day_df['Month Criterion'] = \
            ((yr_mth_day_df['mth']>=1) & (yr_mth_day_df['mth']<=12))
            
        yr_mth_day_df['Year Criterion'] = (
            ((yr_mth_day_df['yr']>=0) & (yr_mth_day_df['yr']<=9999))
            )
        yr_mth_day_df['All Criterion'] = (
            yr_mth_day_df[['Day Criterion','Month Criterion','Year Criterion']]
            .all(axis=1))
        
        if not yr_mth_day_df['All Criterion'].all():
            msg = ('Date series is not "{}", where '
                   'd = Day, m = Month, y = Year.'
                   .format(date_format.lower().strip()))
            # logger.error(msg)
            raise Exception(msg)
            
    elif pd.api.types.is_datetime64_any_dtype(date_series):
        msg = ('Already a datetime format').format(date_series.name)
        # logger.info(msg)
        
    else:
        msg = ("Invalid input type: {}. Please convert to pandas.Series."
               .format(date_series.dtype)
               )
        # logger.error(msg)
        raise NotImplementedError(msg)

### Class
# Parent Class
class ExcelHelper:
                
    def xlref(self, row, column, negative = False):   
        """
        xlref - Simple conversion of row, column to an excel string 
        format
    
        >>> xlref(1,1)
        'A1'
        >>> xlref(1,27, negative=True)
        '-AA1'
        """
                
        cell = utils.get_column_letter(column) + str(row)
        
        if negative is True:

            cell = '-' + cell             
        
        return cell
    
    def addition_formula(self, value_list, equal_sign = True, negative = False):
        """
        Creates a string denoting an Excel formula whcich sums ups 
        all items in the value_list.
        
        >>> addition_formula(['A1', 'B1'])
        '=A1+B1'
        
        >>> addition_formula(["A1", "B1"], equal_sign = True, negative=True)
        '=-(A1+B1)'
        """
        
        # Calculate the terms and cover with brackets
        formula = "(" + "+".join(map(str, value_list)) + ")"
        
        # add negative sign if necessary
        if negative is True:
            formula = "-" + formula
        
        # add equal sign if needed
        if equal_sign is True:
            
            formula = "=" + formula
            
        return formula

    
    def subtraction_formula(self, 
                            value_list, 
                            equal_sign = True, negative = False):
        """
        Creates a string denoting an Excel formula whcich subtracts 
        all items from value_list, except the first, from the first 
        item of value_list.
        
        >>> substraction_formula(['A1', 'B1'])
        '=A1-B1'
        
        >>> subtraction_formula(["A1", "A2", "B3"], True, True)
        '=-(A1-A2-B3)'
        """
        
        # Calculate the terms and cover with brackets
        formula = "(" + "-".join(map(str, value_list)) + ")"
        
        # add negative sign if necessary
        if negative is True:
            formula = "-" + formula
        
        # add equal sign if needed
        if equal_sign is True:
            
            formula = "=" + formula
            
        return formula

    
    def divide_formula(self, exp1, exp2, 
                       equal_sign = True, negative = False):
        """
        Creates a string denoting an Excel formula whcich divides 
        the 2nd expression (exp2) by the first (exp1).
        
        >>> divide_formula('A1-B1', 'C1+D1')
        '=(A1-B1)/(C1+D1)'
        """
        formula = "(" + str(exp1) + ")" + "/" + "(" + str(exp2) + ")"
        
        if negative is True:
            formula = "-" + formula
            
        if equal_sign is True:
            formula = "=" + formula
        
        return formula
    
    def multiply_formula(self, exp1, exp2, 
                         equal_sign = True, negative = False):
        """
        Creates a string denoting an Excel formula whcich multiplies 
        the 1st and 2nd expression together (exp1 and exp2 
        respectively).
        
        >>> multiply_formula('A1-B1', 'C1+D1')
        '=(A1-B1)*(C1+D1)'
        """
        
        formula = "(" + str(exp1) + ")" + "*" + "(" + str(exp2) + ")"
        
        if negative is True:
            formula = "-" + formula
            
        if equal_sign is True:
            formula = "=" + formula

        return formula
    
    def sum_formula(self, 
                    range_start, range_end, 
                    equal_sign = True, negative = False):
        """
        Creates a string denoting an Excel SUM formula for a range 
        of cells.
        
        >>> sum_formula('A1', 'A99')
        '=SUM(A1:A99)'
        """
        
        formula = "SUM(" + str(range_start) + ":" + str(range_end) + ")"
        
        if negative is True:
            formula = "-" + formula
            
        if equal_sign is True:
            formula = "=" + formula
            
        return formula
    
    def count_formula(self, 
                      range_start, range_end, 
                      equal_sign = True, negative = False):
        """
        Creates a string denoting an Excel COUNT formula for a range 
        of cells.
        
        >>> count_formula('A1', 'A99')
        '=COUNT(A1:A99)'
        """
        
        formula = "COUNT(" + str(range_start) + ":" + str(range_end) + ")"
        
        if negative is True:
            formula = "-" + formula
            
        if equal_sign is True:
            formula = "=" + formula
            
        return formula
    
    def replace_cell_value(self, ws, cell_loc, from_str, to_str):
        '''
        Inplace replacing string in excel cell.
        cell_loc is "A1" etc
        
        Implements:
            ws[cell_loc].value = ws[cell_loc].value.replace(from_str, to_str)
        
        #CHANGES:
        #20210207 - added
        '''
        
        original_value = ws[cell_loc].value
        new_value = original_value.replace(from_str, to_str)
        ws[cell_loc].value = new_value
    
    def round_formula(self, 
                      number, num_digits, 
                      equal_sign = True, negative = False):
        """
        Creates a string denoting a Excel ROUND formula for a number
        up to the specified number of digits (num_digits).
        
        >>> round_formula('A1', 2)
        '=ROUND(A1,2)'
        """
        formula = "ROUND(" + str(number) + "," + str(num_digits) + ")"
        
        if negative is True:
            formula = "-" + formula
            
        if equal_sign is True:
            formula = "=" + formula

        return formula
    
    def if_formula(self, 
                   condition, true_result, false_result, 
                   equal_sign = True, negative = False):
        """
        Creates a string denoting an Excel IF formula for a condition
        and 2 results for whether the condition is true or false.
        
        >>> if_formula('A1>2', 2, 3)
        '=IF(A1>2,2,3)'
        """
        
        formula = "IF(" + str(condition) + "," + str(true_result) \
                  + "," + str(false_result) + ")"
        
        if negative is True:
            formula = "-" + formula
            
        if equal_sign is True:
            formula = "=" + formula

        return formula
    
    def isblank_formula(self, cell, equal_sign = True):
        """
        Creates a string denoting a Excel ISBLANK formula for a cell.
        
        >>> isblank_formula('A1')
        '=ISBLANK(A1)'
        """
        
        formula = "ISBLANK(" + str(cell) + ")"
        if equal_sign is True:
            formula = "=" + formula
        
        return formula
    
    def ws_reference(self, ws_name, equal_sign = True):
        """
        Creates a string denoting a worksheet reference for a cell.
        
        >>> ws_reference('test_sheet')
        '=test_sheet!'
        """
        if equal_sign is True:
            formula = "='" + ws_name + "'!"
            
        elif equal_sign is False:
            formula = "'" + ws_name + "'!"
            
        else:
            
            msg = f"equal_sign = {equal_sign} (should be boolean)." 
            raise NotImplementedError (msg)
            
        return formula
    
    def format_range(self, 
                     ws, row_col, 
                     new_min_index, new_max_index, 
                     row_diff = -1, col_diff = -1, 
                     insert_new = True):
        """
        Insert new rows or columns and format the cells based on a 
        default formatted row or column in the same sheet.
        
        #GS: Note that indices are base 1, i.e. row 1 in excel = 1
        
        Inputs:
            - ws: worksheet containing cells to be formatted
            - row_col: format either by row or by column
            - new_min_index: index of first row/col of new range
            - new_max_index: index of last row/col of new range
            - row_diff: default (-1) - number of rows between each range
            - col_diff: default (-1) - number of cols between each range
            - insert_new: default (True) - to insert new cells (True) or 
                          format existing cells (False)
        
        Formatting can be applied on several rows or columns of 
        cells that are of the same dimensions and are a fixed 
        distance apart from each other and the default range.
        """
        
        if row_col == "row":
            if insert_new is True:
                
                num_to_insert = new_max_index - new_min_index + 1
                ws.insert_rows(new_min_index, num_to_insert)
            
            for row in ws.iter_rows(min_row = new_min_index, max_row = new_max_index):
                for cell in row:
                    ref_row = cell.row + row_diff
                    ref_col = cell.col_idx + col_diff
                    ref_cell = ws.cell(row = ref_row, column = ref_col)
                    if ref_cell.has_style:
                        cell.font = copy(ref_cell.font)
                        cell.border = copy(ref_cell.border)
                        cell.fill = copy(ref_cell.fill)
                        cell.number_format = copy(ref_cell.number_format)
                        cell.protection = copy(ref_cell.protection)
                        cell.alignment = copy(ref_cell.alignment)
        
        elif row_col == "col":
            if insert_new is True:
                ws.insert_cols(new_min_index, new_max_index - new_min_index + 1)
            
            for col in ws.iter_cols(min_col = new_min_index, max_col = new_max_index):
                for cell in col:
                    ref_cell = ws.cell(row = cell.row + row_diff, column = cell.col_idx + col_diff)
                    if ref_cell.has_style:
                        cell.font = copy(ref_cell.font)
                        cell.border = copy(ref_cell.border)
                        cell.fill = copy(ref_cell.fill)
                        cell.number_format = copy(ref_cell.number_format)
                        cell.protection = copy(ref_cell.protection)
                        cell.alignment = copy(ref_cell.alignment)
        
        else:
            msg = f"row_col = {row_col}. Should be 'row' or 'col'."
            raise NotImplementedError (msg)                
        return ws
    
    def reorder_sheets(self, wb, ws, new_idx_pos):
        """
        Reorder sheets in an excel workbook.
        """
        old_idx_pos = wb.worksheets.index(ws)
        sheets = wb._sheets.copy()
        sheets.insert(new_idx_pos, sheets.pop(old_idx_pos))
        wb._sheets = sheets
        
        return wb
    
    def df_to_worksheet(self, df, ws, 
                    index=True, header=True,
                    startrow=1, startcol=1
                    ):
        '''
        This function writes a dataframe to an openpyxl worksheet.
        
        This is analogous to the pd.DataFrame.to_excel, which takes
        in a pandas ExcelWriter rather than a worksheet directly.
        
        Inputs:
            - df: A pandas dataframe
            - ws: An openpyxl worksheet object
            - index: bool (default True) - write the index
            - header: bool (default True) - write the header
            - startrow: default (1) - base 1 index where 1 = first row
            - startcol: default (1) - base 1 index where 1 = column A
        
        -----------------------------------------------------------
        # Sample usage
        > wb = openpyxl.Workbook()
        > ws = wb.create_sheet()
        
        > df = pd.DataFrame([[1,2,3], [4,5,6], [7,8,9]], index=list("ABC"),
                          columns=["C1", "C2", "C3"])
            
        > df_to_worksheet(df, ws, index=False, header=False,
                        startrow=3, startcol=3)
        
        >wb.save("test to worksheet.xlsx")
        
        ----------------------------------------------------------
        
        CHANGELOGS:
        20200817 - initialised by owgs
        '''
    
        # Prepare the dataframe based on whether need to write index or header
        if index is True:
            df = df.reset_index()
            
        if header is True:
            df = df.T.reset_index().T
        
        # Get the number of rows and columns to write
        num_rows, num_cols = df.shape
        
        # Set the start row and cols
        for r_offset in range(num_rows):
            
            ridx = startrow + r_offset
            
            for c_offset in range(num_cols):
                
                value = df.iat[r_offset, c_offset]
        
                # Get the cell
                cidx = startcol + c_offset
                c_alpha = utils.get_column_letter(cidx)
                cell = "%s%s" % (c_alpha, ridx)
                
                # Set the value
                ws[cell] = value
                

    def get_indices(self, dfObj, value):
        ''' 
        Get index positions of value in dataframe i.e. dfObj.
        From: https://thispointer.com/python-find-indexes-of-an-element-in-pandas-dataframe/
        
        This returns the a list of positions such as:
            - [(row1, col1), (row2, col2), ... ], where row and col are 
              base0 indices.
        '''
        listOfPos = list()
        # Get bool dataframe with True at positions where the given value exists
        result = dfObj.isin([value])
        # Get list of columns that contains the value
        seriesObj = result.any()
        columnNames = list(seriesObj[seriesObj == True].index)
        # Iterate over list of columns and fetch the rows indexes where value exists
        for col in columnNames:
            rows = list(result[col][result[col] == True].index)
            for row in rows:
                listOfPos.append((row, col))
        # Return a list of tuples indicating the positions of value in the dataframe
        return listOfPos


def check_filepath(fp):
    '''
    This function returns a filepath that does not exist yet.
    
    This is typically used when you want to generate an output
    file, but do not wish to overwrite any file.
    
    Therefore, this function ensures that if the desired fp 
    is already present, then it will add a datetimestamp
    to force a new fp.
    
    Otherwise, it will return the fp that was provided to this 
    function.
    '''
    
    if not os.path.exists(fp):
               
        return_fp = fp

    else:
        
        # Get the components
        dirname, file_name_ext = os.path.split(fp)
        
        # Get the folderpath
        if dirname == "":
            dirname = os.getcwd()
                    
        # get the filename and extension
        file_name, file_ext = os.path.splitext(file_name_ext)
        
        while True:
            
            # Get the current datetimestamp
            datetimestamp = time.strftime("%Y%m%d%H%M%S")
            
            # create the new file
            new_file_name_ext = file_name + "_" + datetimestamp + file_ext
            
            new_fp = os.path.join(dirname, new_file_name_ext)
            
            if os.path.exists(new_fp):
                
                pass
            
            else:
                
                return_fp = new_fp
                break
            
    return return_fp    
        

def add_file_prefix(fp, prefix="", ensure_new_fp = True):
    '''
    Adds a prefix to the filename.
    
    if ensure_new_fp is True, this will ensure that a new filepath
    if returned.
    '''
    
    # Get the components
    dirname, file_name_ext = os.path.split(fp)
    
    # set the new file name
    new_file_name_ext = prefix + file_name_ext
    
    # get the return fp
    new_fp = os.path.join(dirname, new_file_name_ext)
    
    # Check if need to be unique
    if ensure_new_fp:
        
        new_fp = check_filepath(new_fp)
        
    return new_fp

class Constants:
    
    def __init__(self, fp, sheet_name = "CONSTANTS"):
        
        df = pd.read_excel(fp, sheet_name = sheet_name, index_col = 0,
                           header = 0)
        df.columns = df.columns.map(lambda v: v.lower()[0])       
        
        self.df = df
        
    def get(self, var, row_col):
        
        row_col = row_col.lower()[0]
        
        return int(self.df.at[var, row_col])
    
    def getr(self, var):
        
        return int(self.df.at[var, 'r'])
    
    def getc(self, var):
        
        return int(self.df.at[var, 'c'])
    
    def get_loc(self, var):
        
        return (self.getr(var), self.getc(var)) 


class YearCategorisation:
    
    def __init__(self, year_ticks, categories = ["min_pay", "fin_charge", "pv"]):
        '''
        This is a class that splits the period into year bins
        based on the supplied year_ticks.
        
        For example, if year_ticks is [1,2,3], then the year data will
        be split into the following bins:
            - less than 1 year
            - 1 to 2 year
            - 2 to 3 year
            - more than 3 years
        '''
        
        self.year_ticks = year_ticks
        self.categories = categories
        self.year_cut_df = self.categorise_years()
        
    def categorise_years(self):
        
        # Get
        year_ticks = self.year_ticks
        categories = self.categories
        
        # Check the num of ticks
        len_ticks = len(year_ticks)
        #assert len_ticks > 1, "Year ticks must be at least 2."
        
        # cut
        year_cut_df = pd.DataFrame(columns = categories + ['english', 'bounds'])
        
        # Loop
        for i, yr in enumerate(year_ticks, 1):
            
            if i == 1:
                
                prev_yr = 0
                gen_name = f"less_{yr}y"
                gen_name_eng = f"Not later than {yr} year"
            
            else:
                
                prev_yr = year_ticks[i-2]
                gen_name = f"{prev_yr}y_to_{yr}y"
                gen_name_eng = f"Between {prev_yr} to {yr} years"
    
            # Loop thru the cateories
            for cat in categories:
                
                year_cut_df.at[gen_name, cat] = f"{cat}_{gen_name}"
                
            # Set the english name bounds
            year_cut_df.at[gen_name, 'english'] = gen_name_eng
            year_cut_df.at[gen_name, "bounds"] = tuple([prev_yr, yr])
                
        # For last
        gen_name = f"more_{year_ticks[-1]}y"
        gen_name_eng = f"Later than {year_ticks[-1]} years"
        for cat in categories:
            
            year_cut_df.at[gen_name, cat] = f"{cat}_{gen_name}"
            
        # Set the bounds
        year_cut_df.at[gen_name, "bounds"] = tuple([year_ticks[-1], 999999])
        year_cut_df.at[gen_name, "english"] = gen_name_eng
            
        return year_cut_df
    
    
#%%  function tester
if __name__ == "__main__":
    
    if False:
        
        self = YearCategorisation([1])
        year_cut_df = self.year_cut_df
    
    if False:
        
        fp = 'out.txt'
        add_file_prefix(fp, 'test_')
        
    if False:
        
        # Try by changing fp to a file that is already present on your computer.
        fp = 'out.txt'
        check_filepath(fp)
    
    if False:
        # will return error	
        get_unique_dict_values({1: ['a', 'b'], 2: ['c', 'd']}, assert_one_value = True)
    
    if False:
                
        date = '1/12/2021'
        date = '2/12/2021'
        date = '30/12/2021'
        date = '31/12/2021'
        date = '30/11/2021'
    
        shift_month(date, 12, month_begin = True)
        date = shift_month(date, -1)

    #%% format range tester
    if True:
        fn = 'test2.xlsx'
        self = ExcelHelper()
        
        wb = openpyxl.load_workbook(fn)
        ws = wb.copy_worksheet(wb['lead_template'])
            
        ws.insert_rows(33, 5)

        row_col = "row"
        new_min_index = 33
        new_max_index = 35
        row_diff = -1
        col_diff = 0
        insert_new = True
        
        #self.format_range(ws, 'row', 
        #                  new_min_index, new_max_index, 
        #                  row_diff = row_diff, col_diff = col_diff,
        #                  insert_new = True)
        wb.save('book1_output.xlsx')

